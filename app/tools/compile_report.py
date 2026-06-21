import json
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any


def compile_report(run_id: str) -> dict:
    """
    Generate the final artifacts for the Reporter agent:
      - report.html: modern browser report
      - report.pdf: printable/shareable report, when WeasyPrint is available
      - report.xlsx: Excel summary for tables and KPI handoff
    """
    try:
        base = Path("runs") / run_id
        artifacts_dir = base / "artifacts"
        artifacts_dir.mkdir(parents=True, exist_ok=True)

        metadata = _read_json(base / "metadata.json", {})
        insights = _read_json(artifacts_dir / "insights.json", {})
        profile = _read_json(artifacts_dir / "profile.json", {})
        cleaning_rules = _read_json(artifacts_dir / "cleaning_rules.json", {})
        dashboard_manifest = _read_json(
            artifacts_dir / "dashboard_artifacts_manifest.json", {}
        )
        bi_handoff = _read_json(artifacts_dir / "bi_agent_handoff.json", {})
        tool_calls = _read_jsonl(base / "tool_calls.jsonl")
        decisions = _read_jsonl(base / "decisions.jsonl")

        kpis = insights.get("kpis", {}) or {}
        alertes = insights.get("alertes", []) or []
        insight_lines = _clean_insight_lines(insights.get("insights", []) or [])
        charts = _collect_charts(bi_handoff, dashboard_manifest, artifacts_dir)
        dashboard_path = (
            bi_handoff.get("dashboard_path")
            or dashboard_manifest.get("dashboard_path")
            or str(artifacts_dir / "dashboard.html")
        )

        report_data = {
            "run_id": run_id,
            "metadata": metadata,
            "insights": insights,
            "profile": profile,
            "cleaning_rules": cleaning_rules,
            "bi_handoff": bi_handoff,
            "dashboard_path": dashboard_path,
            "charts": charts,
            "kpis": kpis,
            "alertes": alertes,
            "insight_lines": insight_lines,
            "decisions": decisions,
            "tool_calls": tool_calls,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        html = _render_html(**report_data)
        html_path = artifacts_dir / "report.html"
        html_path.write_text(html, encoding="utf-8")

        pdf_path, pdf_status = _write_pdf(
            html_path,
            artifacts_dir / "report.pdf",
            report_data,
        )
        excel_path = _write_excel(report_data, artifacts_dir / "report.xlsx")

        return {
            "status": "success",
            "report_path": str(html_path),
            "html_path": str(html_path),
            "pdf_path": str(pdf_path) if pdf_path else "",
            "pdf_status": pdf_status,
            "excel_path": str(excel_path),
            "dashboard_path": dashboard_path,
            "run_id": run_id,
            "nb_kpis": len(kpis),
            "nb_alertes": len(alertes),
            "nb_charts": len(charts),
            "generated": True,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        }

    except Exception as exc:
        return {"status": "error", "error": str(exc), "generated": False}


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def _read_jsonl(path: Path) -> list[dict]:
    rows: list[dict] = []
    if not path.exists():
        return rows
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            rows.append({"raw": line})
    return rows


def _clean_insight_lines(lines: list[Any]) -> list[str]:
    cleaned: list[str] = []
    for line in lines:
        text = str(line).strip()
        if not text or set(text) <= {"=", "-", "_"}:
            continue
        cleaned.append(text)
    return cleaned[:10]


def _collect_charts(
    bi_handoff: dict,
    dashboard_manifest: dict,
    artifacts_dir: Path,
) -> list[str]:
    charts: dict[str, str] = {}
    for item in bi_handoff.get("charts", []) or []:
        if item:
            clean = _clean_path(item)
            charts[clean.lower()] = clean
    for item in dashboard_manifest.get("charts", []) or []:
        path = item.get("chart_path") or item.get("path") if isinstance(item, dict) else item
        if path:
            clean = _clean_path(path)
            charts[clean.lower()] = clean

    charts_dir = artifacts_dir / "charts"
    if charts_dir.exists():
        for path in charts_dir.glob("*.html"):
            clean = _clean_path(path)
            charts[clean.lower()] = clean

    return list(charts.values())


def _render_html(
    run_id: str,
    metadata: dict,
    insights: dict,
    profile: dict,
    cleaning_rules: dict,
    bi_handoff: dict,
    dashboard_path: str,
    charts: list[str],
    kpis: dict,
    alertes: list,
    insight_lines: list[str],
    decisions: list[dict],
    tool_calls: list[dict],
    generated_at: str,
) -> str:
    quality = insights.get("data_quality", {}) or {}
    headline_metrics = _headline_metrics(kpis, bi_handoff)
    artifacts = metadata.get("artifacts", {}) or {}
    domain = insights.get("domain") or kpis.get("domain_detected") or "unknown"
    objective = metadata.get("objective") or "Non renseigne"
    status = metadata.get("status") or "completed"
    quality_score = quality.get("score_global") or kpis.get("data_quality_score")

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Analytics Report - {escape(run_id)}</title>
  <style>
    @page {{ size: A4; margin: 14mm; }}
    :root {{
      --bg: #f4f7fb;
      --surface: #ffffff;
      --ink: #172033;
      --muted: #667085;
      --line: #d8e0eb;
      --blue: #155eef;
      --blue-soft: #eaf1ff;
      --teal: #0f766e;
      --amber: #b45309;
      --red: #b42318;
      --shadow: 0 18px 45px rgba(31, 42, 68, 0.10);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font-family: Inter, Segoe UI, Arial, sans-serif;
      line-height: 1.5;
    }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 28px 18px 56px; }}
    .hero {{
      color: #fff;
      border-radius: 22px;
      padding: 30px;
      background:
        linear-gradient(135deg, rgba(21,94,239,.96), rgba(15,118,110,.94)),
        linear-gradient(45deg, #155eef, #0f766e);
      box-shadow: var(--shadow);
      margin-bottom: 20px;
    }}
    .hero-top {{ display: flex; justify-content: space-between; gap: 18px; align-items: flex-start; }}
    .eyebrow {{ margin: 0 0 8px; text-transform: uppercase; letter-spacing: .08em; font-size: 12px; opacity: .82; }}
    h1 {{ margin: 0; font-size: 34px; line-height: 1.12; }}
    .hero p {{ margin: 10px 0 0; color: rgba(255,255,255,.86); }}
    .badge {{ background: rgba(255,255,255,.16); border: 1px solid rgba(255,255,255,.28); border-radius: 999px; padding: 8px 12px; white-space: nowrap; }}
    .cards {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 12px; margin-top: 22px; }}
    .metric {{
      background: rgba(255,255,255,.14);
      border: 1px solid rgba(255,255,255,.22);
      border-radius: 16px;
      padding: 14px;
      min-height: 92px;
    }}
    .metric .label {{ color: rgba(255,255,255,.74); font-size: 11px; text-transform: uppercase; }}
    .metric .value {{ margin-top: 6px; font-size: 22px; font-weight: 760; overflow-wrap: anywhere; }}
    section {{
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 18px;
      padding: 20px;
      margin: 16px 0;
      box-shadow: 0 10px 28px rgba(31, 42, 68, 0.055);
    }}
    .section-head {{ display: flex; align-items: center; justify-content: space-between; gap: 14px; margin-bottom: 14px; }}
    h2 {{ margin: 0; font-size: 19px; color: #0f2f6b; }}
    .tag {{ display: inline-block; border-radius: 999px; background: var(--blue-soft); color: #174ea6; padding: 5px 10px; font-size: 12px; font-weight: 700; }}
    table {{ width: 100%; border-collapse: separate; border-spacing: 0; overflow: hidden; border: 1px solid var(--line); border-radius: 12px; }}
    th, td {{ padding: 10px 12px; border-bottom: 1px solid var(--line); text-align: left; vertical-align: top; }}
    th {{ background: #f0f5ff; color: #344054; font-size: 12px; text-transform: uppercase; }}
    tr:last-child td {{ border-bottom: 0; }}
    td:last-child {{ overflow-wrap: anywhere; }}
    .two-col {{ display: grid; grid-template-columns: 1.05fr .95fr; gap: 16px; }}
    .insights {{ display: grid; gap: 10px; }}
    .insight {{ border-left: 4px solid var(--blue); background: #f8fbff; padding: 11px 12px; border-radius: 10px; }}
    .alert {{ border-left: 4px solid var(--amber); background: #fffbeb; padding: 11px 12px; border-radius: 10px; margin: 8px 0; }}
    .alert.critical, .alert.danger {{ border-left-color: var(--red); background: #fff1f3; }}
    .empty {{ color: var(--muted); font-style: italic; }}
    .path {{ color: var(--teal); font-family: Consolas, monospace; font-size: 13px; overflow-wrap: anywhere; }}
    .flow {{ display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }}
    .flow span {{ background: #f0f5ff; border: 1px solid #d8e6ff; color: #164496; border-radius: 999px; padding: 7px 10px; font-weight: 700; font-size: 13px; }}
    .flow b {{ color: var(--muted); }}
    .footer {{ color: var(--muted); text-align: center; font-size: 12px; margin-top: 24px; }}
    @media print {{
      body {{ background: #fff; }}
      main {{ padding: 0; }}
      section, .hero {{ box-shadow: none; break-inside: avoid; }}
    }}
    @media (max-width: 900px) {{
      .cards {{ grid-template-columns: repeat(2, 1fr); }}
      .two-col {{ grid-template-columns: 1fr; }}
      .hero-top {{ flex-direction: column; }}
    }}
    @media (max-width: 560px) {{
      main {{ padding: 14px 10px 34px; }}
      .cards {{ grid-template-columns: 1fr; }}
      .hero {{ border-radius: 16px; padding: 22px; }}
      h1 {{ font-size: 27px; }}
    }}
  </style>
</head>
<body>
<main>
  <div class="hero">
    <div class="hero-top">
      <div>
        <p class="eyebrow">AI multi-agent analytics ecosystem</p>
        <h1>Rapport final analytics</h1>
        <p>Run {escape(run_id)} - genere le {escape(generated_at)}</p>
      </div>
      <div class="badge">{escape(str(status)).upper()}</div>
    </div>
    <div class="cards">
      {_hero_metric("Domaine", domain)}
      {_hero_metric("Score qualite", _format_percent(quality_score))}
      {_hero_metric("KPIs", len(kpis))}
      {_hero_metric("Alertes", len(alertes))}
    </div>
  </div>

  <section>
    <div class="section-head"><h2>Resume executif</h2><span class="tag">overview</span></div>
    <p><strong>Objectif:</strong> {escape(str(objective))}</p>
    {_headline_cards(headline_metrics)}
  </section>

  <div class="two-col">
    <section>
      <div class="section-head"><h2>KPIs business</h2><span class="tag">{len(kpis)} indicateurs</span></div>
      {_kpi_table(kpis, headline_metrics)}
    </section>

    <section>
      <div class="section-head"><h2>Data quality</h2><span class="tag">controle</span></div>
      {_quality_block(quality, profile, cleaning_rules)}
    </section>
  </div>

  <section>
    <div class="section-head"><h2>Insights principaux</h2><span class="tag">narrative</span></div>
    {_insights_block(insight_lines)}
  </section>

  <section>
    <div class="section-head"><h2>Alertes</h2><span class="tag">{len(alertes)} detectee(s)</span></div>
    {_alerts_block(alertes)}
  </section>

  <section>
    <div class="section-head"><h2>Dashboard et artifacts</h2><span class="tag">livrables</span></div>
    {_artifacts_table(artifacts, dashboard_path, charts)}
  </section>

  <section>
    <div class="section-head"><h2>Communication des agents</h2><span class="tag">workflow</span></div>
    {_agent_flow_block(bi_handoff, decisions)}
  </section>

  <section>
    <div class="section-head"><h2>Tools MCP utilises</h2><span class="tag">audit</span></div>
    {_tool_calls_table(tool_calls)}
  </section>

  <p class="footer">Generated by Reporter Agent - {escape(run_id)}</p>
</main>
</body>
</html>"""


def _headline_metrics(kpis: dict, bi_handoff: dict) -> dict:
    raw = bi_handoff.get("headline_metrics", {}) or {}
    if isinstance(raw, list):
        raw = {item["key"]: item["value"] for item in raw if isinstance(item, dict) and "key" in item and "value" in item}
    cleaned = {key: value for key, value in raw.items() if value not in (None, "", "N/A")}

    aliases = {
        "Revenue": ["Sales Revenue", "CA_total", "total_revenue", "revenue", "Average Order Value"],
        "Orders": ["Order Count", "nb_commandes", "orders", "total_orders"],
        "Customers": ["Customer Count", "nb_clients_uniques", "customers", "unique_customers"],
        "Return rate": ["Return Rate", "taux_annulation", "return_rate"],
    }
    for label, keys in aliases.items():
        if label in cleaned:
            continue
        for key in keys:
            if key in kpis and kpis[key] not in (None, "", "N/A"):
                cleaned[label] = kpis[key]
                break

    if not cleaned:
        for key, value in kpis.items():
            if isinstance(value, (int, float, str, dict)) and key != "domain_detected":
                cleaned[key] = value
            if len(cleaned) >= 4:
                break

    return dict(list(cleaned.items())[:4])


def _hero_metric(label: str, value: Any) -> str:
    return (
        '<div class="metric">'
        f'<div class="label">{escape(str(label))}</div>'
        f'<div class="value">{escape(_format_value(value))}</div>'
        '</div>'
    )


def _headline_cards(metrics: dict) -> str:
    if not metrics:
        return '<p class="empty">Aucun indicateur de synthese disponible.</p>'
    return '<div class="cards">' + "".join(
        _hero_metric(key, value) for key, value in metrics.items()
    ) + "</div>"


def _artifacts_table(artifacts: dict, dashboard_path: str, charts: list[str]) -> str:
    report_base = _clean_path(dashboard_path).replace("dashboard.html", "")
    rows = [
        ("Clean CSV", _clean_path(artifacts.get("clean_csv"))),
        ("Insights JSON", _clean_path(artifacts.get("insights_json"))),
        ("Dashboard HTML", _clean_path(dashboard_path)),
        ("Charts", f"{len(charts)} chart(s): " + ", ".join(_clean_path(c) for c in charts[:4])),
        ("Report HTML", _clean_path(artifacts.get("report_path") or f"{report_base}report.html")),
        ("Report PDF", f"{report_base}report.pdf"),
        ("Report Excel", f"{report_base}report.xlsx"),
    ]
    rows = [(name, value) for name, value in rows if value not in ("", "N/A")]
    return _simple_table(["Artifact", "Valeur"], rows)


def _kpi_table(kpis: dict, headline_metrics: dict) -> str:
    merged = {**headline_metrics, **kpis}
    merged = {
        key: value
        for key, value in merged.items()
        if value not in (None, "", "N/A") and not _is_empty_headline_alias(key, value)
    }
    if not merged:
        return '<p class="empty">Aucun KPI disponible.</p>'
    rows = [(key, _format_value(value)) for key, value in merged.items()]
    return _simple_table(["KPI", "Valeur"], rows)


def _quality_block(quality: dict, profile: dict, cleaning_rules: dict) -> str:
    rows = [
        ("Score global", _format_percent(quality.get("score_global"))),
        ("Missing rate", _format_percent(quality.get("missing_rate"))),
        ("Lignes", _first_present(quality.get("nb_lignes"), profile.get("rows"))),
        ("Colonnes", _first_present(quality.get("nb_colonnes"), profile.get("columns"))),
        ("Doublons", _first_present(quality.get("nb_doublons"), profile.get("duplicate_rows"), 0)),
        ("Regles nettoyage", len(cleaning_rules) if isinstance(cleaning_rules, dict) else 0),
    ]
    return _simple_table(["Element", "Valeur"], rows)


def _insights_block(lines: list[str]) -> str:
    if not lines:
        return '<p class="empty">Aucun insight disponible.</p>'
    return '<div class="insights">' + "".join(
        f'<div class="insight">{escape(line)}</div>' for line in lines
    ) + "</div>"


def _alerts_block(alertes: list) -> str:
    if not alertes:
        return '<p class="empty">Aucune alerte detectee.</p>'
    html = []
    for alert in alertes:
        if isinstance(alert, dict):
            level = str(alert.get("niveau") or alert.get("level") or "warning")
            kpi = alert.get("kpi", "KPI")
            message = alert.get("message", "")
            value = alert.get("valeur", "")
        else:
            level = "warning"
            kpi = "Alerte"
            message = str(alert)
            value = ""
        html.append(
            f'<div class="alert {escape(level)}">'
            f'<strong>{escape(str(level)).upper()} - {escape(str(kpi))}</strong><br>'
            f'{escape(_format_value(value))} {escape(str(message))}'
            '</div>'
        )
    return "".join(html)


def _agent_flow_block(bi_handoff: dict, decisions: list[dict]) -> str:
    flow = bi_handoff.get("agent_context", {}).get("flow") or [
        "Data Engineer",
        "Data Scientist",
        "BI Agent",
        "Reporter",
    ]
    html = '<div class="flow">'
    for idx, item in enumerate(flow):
        if idx:
            html += "<b>to</b>"
        html += f"<span>{escape(str(item))}</span>"
    html += "</div>"

    if decisions:
        rows = [
            (
                item.get("timestamp", "")[:19],
                item.get("agent", ""),
                item.get("decision", ""),
                item.get("reason", ""),
            )
            for item in decisions[-15:]
        ]
        html += "<br>" + _simple_table(["Timestamp", "Agent", "Decision", "Raison"], rows)
    return html


def _tool_calls_table(tool_calls: list[dict]) -> str:
    if not tool_calls:
        return '<p class="empty">Aucun log tool_calls.jsonl trouve.</p>'
    summarized: dict[tuple[str, str], dict] = {}
    for call in tool_calls:
        agent = call.get("agent_name") or call.get("agent") or ""
        tool = call.get("tool_name") or call.get("tool") or ""
        if not agent or agent == "N/A" or not tool:
            continue
        key = (agent, tool)
        item = summarized.setdefault(key, {
            "timestamp": "",
            "agent": agent,
            "tool": tool,
            "count": 0,
            "success": True,
            "failures": 0,
        })
        item["timestamp"] = call.get("timestamp", "")[:19] or item["timestamp"]
        item["count"] += 1
        item["success"] = bool(call.get("success", True))
        if not bool(call.get("success", True)):
            item["failures"] += 1

    rows = []
    for call in list(summarized.values())[-20:]:
        status = "OK" if call["success"] else "ERROR"
        if call["failures"] and call["success"]:
            status = f"OK ({call['failures']} ancien(s) echec(s))"
        rows.append(
            (
                call["timestamp"],
                call["agent"],
                call["tool"],
                call["count"],
                status,
            )
        )
    if not rows:
        return '<p class="empty">Aucun appel MCP exploitable trouve.</p>'
    return _simple_table(["Dernier appel", "Agent", "Tool", "Nb appels", "Statut"], rows)


def _simple_table(headers: list[str], rows: list[tuple[Any, ...]]) -> str:
    head = "".join(f"<th>{escape(str(header))}</th>" for header in headers)
    body = []
    for row in rows:
        body.append(
            "<tr>"
            + "".join(f"<td>{escape(_format_value(cell))}</td>" for cell in row)
            + "</tr>"
        )
    return f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(body)}</tbody></table>"


def _clean_path(value: Any) -> str:
    if value in (None, "", "N/A"):
        return ""
    return str(value).replace("\\", "/")


def _is_empty_headline_alias(key: str, value: Any) -> bool:
    canonical_empty_names = {
        "CA_total",
        "nb_commandes",
        "nb_clients_uniques",
        "panier_moyen",
    }
    return key in canonical_empty_names and value in (None, "", "N/A")


def _first_present(*values: Any) -> Any:
    for value in values:
        if value not in (None, "", "N/A"):
            return value
    return None


def _write_pdf(
    html_path: Path,
    pdf_path: Path,
    report_data: dict,
) -> tuple[Path | None, str]:
    try:
        from weasyprint import HTML

        HTML(filename=str(html_path)).write_pdf(str(pdf_path))
        return pdf_path, "generated"
    except Exception as exc:
        _write_basic_pdf(pdf_path, report_data)
        return pdf_path, f"generated_with_basic_fallback: {exc}"


def _write_basic_pdf(pdf_path: Path, report_data: dict) -> None:
    """Styled dependency-free PDF fallback used when HTML-to-PDF is unavailable."""
    pages = _pdf_report_streams(report_data)
    objects: list[bytes] = []

    def add_object(body: str | bytes) -> int:
        if isinstance(body, str):
            body = body.encode("latin-1", errors="replace")
        objects.append(body)
        return len(objects)

    font_regular_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>")
    font_bold_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >>")
    page_ids = []
    content_ids = []
    for stream in pages:
        content_id = add_object(
            b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n"
            + stream + b"\nendstream"
        )
        content_ids.append(content_id)
        page_ids.append(0)

    pages_id = len(objects) + len(pages) + 1
    for idx, content_id in enumerate(content_ids):
        page_ids[idx] = add_object(
            f"<< /Type /Page /Parent {pages_id} 0 R "
            f"/MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 {font_regular_id} 0 R /F2 {font_bold_id} 0 R >> >> "
            f"/Contents {content_id} 0 R >>"
        )

    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    pages_id = add_object(f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>")
    catalog_id = add_object(f"<< /Type /Catalog /Pages {pages_id} 0 R >>")

    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for idx, body in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{idx} 0 obj\n".encode("ascii"))
        output.extend(body)
        output.extend(b"\nendobj\n")

    xref_pos = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\n"
        f"startxref\n{xref_pos}\n%%EOF\n".encode("ascii")
    )
    pdf_path.write_bytes(bytes(output))


def _pdf_report_streams(report_data: dict) -> list[bytes]:
    width, height = 595, 842
    margin = 42
    page_bottom = 48
    pages: list[list[str]] = []
    page: list[str] = []
    y = height - margin
    page_no = 0

    insights = report_data["insights"]
    quality = insights.get("data_quality", {}) or {}
    kpis = report_data["kpis"]
    headline = _headline_metrics(kpis, report_data["bi_handoff"])
    domain = insights.get("domain") or kpis.get("domain_detected", "unknown")
    status = report_data["metadata"].get("status", "completed")
    quality_score = _format_percent(quality.get("score_global") or kpis.get("data_quality_score"))

    def new_page(title: str = "Rapport final analytics") -> None:
        nonlocal page, y, page_no
        if page:
            _pdf_footer(page, page_no)
            pages.append(page)
        page_no += 1
        page = []
        y = height - margin
        _pdf_rect(page, 0, height - 78, width, 78, (0.94, 0.97, 1.0))
        _pdf_rect(page, 0, height - 78, 8, 78, (0.08, 0.37, 0.94))
        _pdf_text(page, margin, height - 38, title, 16, "F2", (0.05, 0.13, 0.27))
        _pdf_text(
            page,
            margin,
            height - 58,
            f"Run {report_data['run_id']} - {report_data['generated_at']}",
            9,
            "F1",
            (0.39, 0.45, 0.55),
        )
        y = height - 105

    def ensure(space: int) -> None:
        if y - space < page_bottom:
            new_page()

    def section(title: str, tag: str = "") -> None:
        nonlocal y
        ensure(44)
        _pdf_text(page, margin, y, title, 14, "F2", (0.05, 0.18, 0.42))
        if tag:
            _pdf_round_label(page, width - margin - 78, y - 11, 78, 18, tag)
        y -= 18
        _pdf_line(page, margin, y, width - margin, y, (0.85, 0.89, 0.94))
        y -= 16

    def paragraph(text: str, size: int = 10, max_width: int = 96) -> None:
        nonlocal y
        for line in _wrap_text(text, max_width):
            ensure(14)
            _pdf_text(page, margin, y, line, size, "F1", (0.12, 0.16, 0.24))
            y -= size + 4

    def table(headers: list[str], rows: list[tuple[Any, ...]], col_widths: list[int]) -> None:
        nonlocal y
        row_height = 24
        ensure(row_height * 2)
        x = margin
        _pdf_rect(page, margin, y - row_height + 5, sum(col_widths), row_height, (0.92, 0.95, 1.0))
        for idx, header in enumerate(headers):
            _pdf_text(page, x + 6, y - 10, header, 8, "F2", (0.18, 0.25, 0.36))
            x += col_widths[idx]
        y -= row_height
        for row in rows:
            ensure(row_height + 6)
            x = margin
            _pdf_line(page, margin, y + 4, margin + sum(col_widths), y + 4, (0.86, 0.90, 0.95))
            for idx, cell in enumerate(row):
                max_chars = max(12, int(col_widths[idx] / 5.5))
                text = _format_value(cell)
                if len(text) > max_chars:
                    text = text[: max_chars - 1] + "."
                _pdf_text(page, x + 6, y - 10, text, 8, "F1", (0.10, 0.14, 0.21))
                x += col_widths[idx]
            y -= row_height
        y -= 8

    new_page()

    # Cover hero
    _pdf_rect(page, margin, y - 118, width - 2 * margin, 118, (0.08, 0.37, 0.94))
    _pdf_rect(page, margin, y - 118, width - 2 * margin, 34, (0.06, 0.46, 0.43))
    _pdf_text(page, margin + 18, y - 28, "AI Multi-Agent Analytics Ecosystem", 10, "F2", (0.84, 0.91, 1.0))
    _pdf_text(page, margin + 18, y - 57, "Rapport final analytics", 24, "F2", (1, 1, 1))
    _pdf_text(page, margin + 18, y - 82, f"Run {report_data['run_id']} - genere le {report_data['generated_at']}", 10, "F1", (0.91, 0.95, 1.0))
    _pdf_text(page, width - margin - 95, y - 31, str(status).upper(), 9, "F2", (1, 1, 1))
    y -= 142

    card_w = (width - 2 * margin - 24) / 4
    card_values = [
        ("Domaine", domain),
        ("Score qualite", quality_score),
        ("KPIs", len(kpis)),
        ("Alertes", len(report_data["alertes"])),
    ]
    for idx, (label, value) in enumerate(card_values):
        x = margin + idx * (card_w + 8)
        _pdf_rect(page, x, y - 58, card_w, 58, (0.98, 0.99, 1.0))
        _pdf_stroke_rect(page, x, y - 58, card_w, 58, (0.80, 0.86, 0.94))
        _pdf_text(page, x + 9, y - 18, label.upper(), 7, "F2", (0.38, 0.44, 0.54))
        _pdf_text(page, x + 9, y - 40, _format_value(value), 15, "F2", (0.05, 0.13, 0.27))
    y -= 82

    section("Resume executif", "overview")
    paragraph(f"Objectif: {report_data['metadata'].get('objective') or 'Non renseigne'}")
    headline_rows = [(key, _format_value(value)) for key, value in headline.items()]
    table(["Indicateur", "Valeur"], headline_rows, [230, 275])

    section("KPIs business", f"{len(kpis)} KPIs")
    kpi_rows = [
        (key, _format_value(value))
        for key, value in {**headline, **kpis}.items()
        if value not in (None, "", "N/A") and not _is_empty_headline_alias(key, value)
    ]
    table(["KPI", "Valeur"], kpi_rows[:18], [230, 275])

    section("Data quality", "controle")
    quality_rows = [
        ("Score global", quality_score),
        ("Missing rate", _format_percent(quality.get("missing_rate"))),
        ("Lignes", _first_present(quality.get("nb_lignes"), report_data["profile"].get("rows"))),
        ("Colonnes", _first_present(quality.get("nb_colonnes"), report_data["profile"].get("columns"))),
        ("Doublons residuels", _first_present(quality.get("nb_doublons"), 0)),
        ("Colonnes sans null", len(quality.get("colonnes_ok", []) or [])),
    ]
    table(["Element", "Valeur"], quality_rows, [230, 275])

    section("Insights principaux", "narrative")
    for item in report_data["insight_lines"][:10]:
        paragraph(f"- {item}", max_width=102)

    section("Alertes", f"{len(report_data['alertes'])} detectee(s)")
    if report_data["alertes"]:
        alert_rows = []
        for alert in report_data["alertes"]:
            if isinstance(alert, dict):
                alert_rows.append((
                    alert.get("niveau") or alert.get("level") or "warning",
                    alert.get("kpi", ""),
                    alert.get("message", ""),
                ))
            else:
                alert_rows.append(("warning", "", str(alert)))
        table(["Niveau", "KPI", "Message"], alert_rows, [90, 140, 275])
    else:
        paragraph("Aucune alerte detectee. Les indicateurs sont dans les seuils attendus.")

    section("Dashboard et artifacts", "livrables")
    artifact_rows = [
        ("Clean CSV", _clean_path(report_data["metadata"].get("artifacts", {}).get("clean_csv"))),
        ("Insights JSON", _clean_path(report_data["metadata"].get("artifacts", {}).get("insights_json"))),
        ("Dashboard HTML", _clean_path(report_data["dashboard_path"])),
        ("Charts", f"{len(report_data['charts'])} chart(s)"),
        ("Report HTML", _clean_path(str(Path('runs') / report_data["run_id"] / 'artifacts' / 'report.html'))),
        ("Report PDF", _clean_path(str(Path('runs') / report_data["run_id"] / 'artifacts' / 'report.pdf'))),
        ("Report Excel", _clean_path(str(Path('runs') / report_data["run_id"] / 'artifacts' / 'report.xlsx'))),
    ]
    table(["Artifact", "Chemin"], artifact_rows, [120, 385])

    section("Communication des agents", "workflow")
    flow = report_data["bi_handoff"].get("agent_context", {}).get("flow") or [
        "Data Engineer", "Data Scientist", "BI Agent", "Reporter"
    ]
    paragraph(" -> ".join(flow), max_width=95)

    section("Tools MCP utilises", "audit")
    tool_rows = _summarized_tool_rows(report_data["tool_calls"])
    table(["Agent", "Tool", "Appels", "Statut"], tool_rows[:16], [135, 150, 70, 150])

    _pdf_footer(page, page_no)
    pages.append(page)
    return ["\n".join(commands).encode("latin-1", errors="replace") for commands in pages]


def _summarized_tool_rows(tool_calls: list[dict]) -> list[tuple[Any, ...]]:
    summarized: dict[tuple[str, str], dict] = {}
    for call in tool_calls:
        agent = call.get("agent_name") or call.get("agent") or ""
        tool = call.get("tool_name") or call.get("tool") or ""
        if not agent or agent == "N/A" or not tool:
            continue
        key = (agent, tool)
        item = summarized.setdefault(key, {"agent": agent, "tool": tool, "count": 0, "success": True, "failures": 0})
        item["count"] += 1
        item["success"] = bool(call.get("success", True))
        if not bool(call.get("success", True)):
            item["failures"] += 1
    rows = []
    for item in summarized.values():
        status = "OK" if item["success"] else "ERROR"
        if item["success"] and item["failures"]:
            status = f"OK ({item['failures']} old failures)"
        rows.append((item["agent"], item["tool"], item["count"], status))
    return rows or [("N/A", "N/A", 0, "No MCP calls")]


def _pdf_text(cmds: list[str], x: float, y: float, text: Any, size: int, font: str, color: tuple[float, float, float]) -> None:
    r, g, b = color
    cmds.append(f"{r:.3f} {g:.3f} {b:.3f} rg")
    cmds.append(f"BT /{font} {size} Tf {x:.2f} {y:.2f} Td ({_pdf_escape(_format_value(text))}) Tj ET")


def _pdf_rect(cmds: list[str], x: float, y: float, w: float, h: float, color: tuple[float, float, float]) -> None:
    r, g, b = color
    cmds.append(f"{r:.3f} {g:.3f} {b:.3f} rg {x:.2f} {y:.2f} {w:.2f} {h:.2f} re f")


def _pdf_stroke_rect(cmds: list[str], x: float, y: float, w: float, h: float, color: tuple[float, float, float]) -> None:
    r, g, b = color
    cmds.append(f"{r:.3f} {g:.3f} {b:.3f} RG 0.8 w {x:.2f} {y:.2f} {w:.2f} {h:.2f} re S")


def _pdf_line(cmds: list[str], x1: float, y1: float, x2: float, y2: float, color: tuple[float, float, float]) -> None:
    r, g, b = color
    cmds.append(f"{r:.3f} {g:.3f} {b:.3f} RG 0.6 w {x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S")


def _pdf_round_label(cmds: list[str], x: float, y: float, w: float, h: float, text: str) -> None:
    _pdf_rect(cmds, x, y, w, h, (0.91, 0.95, 1.0))
    _pdf_text(cmds, x + 8, y + 5, text, 7, "F2", (0.09, 0.28, 0.65))


def _pdf_footer(cmds: list[str], page_no: int) -> None:
    _pdf_line(cmds, 42, 35, 553, 35, (0.86, 0.90, 0.95))
    _pdf_text(cmds, 42, 20, "Generated by Reporter Agent", 8, "F1", (0.42, 0.47, 0.55))
    _pdf_text(cmds, 515, 20, f"Page {page_no}", 8, "F1", (0.42, 0.47, 0.55))


def _wrap_text(text: str, max_chars: int) -> list[str]:
    words = str(text).split()
    lines: list[str] = []
    current = ""
    for word in words:
        if len(current) + len(word) + 1 <= max_chars:
            current = f"{current} {word}".strip()
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines or [""]


def _pdf_escape(text: str) -> str:
    safe = text.encode("latin-1", errors="replace").decode("latin-1")
    return safe.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _write_excel(report_data: dict, excel_path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    quality = report_data["insights"].get("data_quality", {}) or {}
    kpis = report_data["kpis"]
    bi_handoff = report_data["bi_handoff"]
    headline = _headline_metrics(kpis, bi_handoff)

    _write_rows(summary, [
        ("Run ID", report_data["run_id"]),
        ("Generated at", report_data["generated_at"]),
        ("Status", report_data["metadata"].get("status", "completed")),
        ("Domain", report_data["insights"].get("domain") or kpis.get("domain_detected", "unknown")),
        ("Quality score", _format_percent(quality.get("score_global") or kpis.get("data_quality_score"))),
        ("Dashboard", report_data["dashboard_path"]),
        ("Charts", len(report_data["charts"])),
        ("Alerts", len(report_data["alertes"])),
    ])

    ws = wb.create_sheet("Headline KPIs")
    _write_rows(ws, [("KPI", "Value")] + [(k, _format_value(v)) for k, v in headline.items()])

    ws = wb.create_sheet("All KPIs")
    _write_rows(ws, [("KPI", "Value")] + [(k, _format_value(v)) for k, v in kpis.items()])

    ws = wb.create_sheet("Data Quality")
    _write_rows(ws, [
        ("Metric", "Value"),
        ("Score global", _format_percent(quality.get("score_global"))),
        ("Missing rate", _format_percent(quality.get("missing_rate"))),
        ("Rows", quality.get("nb_lignes") or report_data["profile"].get("rows")),
        ("Columns", quality.get("nb_colonnes") or report_data["profile"].get("columns")),
        ("Duplicates", quality.get("nb_doublons") or report_data["profile"].get("duplicate_rows", 0)),
    ])

    ws = wb.create_sheet("Alerts")
    alert_rows = [("Level", "KPI", "Value", "Message")]
    for alert in report_data["alertes"]:
        if isinstance(alert, dict):
            alert_rows.append((
                alert.get("niveau") or alert.get("level") or "warning",
                alert.get("kpi", ""),
                _format_value(alert.get("valeur", "")),
                alert.get("message", ""),
            ))
        else:
            alert_rows.append(("warning", "", "", str(alert)))
    _write_rows(ws, alert_rows)

    ws = wb.create_sheet("Charts")
    _write_rows(ws, [("Chart path",)] + [(chart,) for chart in report_data["charts"]])

    ws = wb.create_sheet("Tool Calls")
    rows = [("Timestamp", "Agent", "Tool", "Success")]
    for call in report_data["tool_calls"]:
        rows.append((
            call.get("timestamp", "")[:19],
            call.get("agent_name") or call.get("agent") or "",
            call.get("tool_name") or call.get("tool") or "",
            call.get("success", True),
        ))
    _write_rows(ws, rows)

    for sheet in wb.worksheets:
        _style_sheet(sheet)
    wb.save(excel_path)
    return excel_path


def _write_rows(sheet, rows: list[tuple[Any, ...]]) -> None:
    for row in rows:
        sheet.append([_format_value(cell) for cell in row])


def _style_sheet(sheet) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="EAF1FF")
    header_font = Font(bold=True, color="173B78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_len + 2, 12), 55)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    sheet.freeze_panes = "A2"


def _format_value(value: Any) -> str:
    if value is None or value == "":
        return "N/A"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        if 0 <= value <= 1:
            return f"{value:.2%}"
        return f"{value:,.2f}"
    if isinstance(value, dict):
        if "total" in value:
            return _format_value(value.get("total"))
        preview = ", ".join(
            f"{key}: {_format_value(val)}"
            for key, val in list(value.items())[:5]
        )
        return preview or "N/A"
    if isinstance(value, list):
        return ", ".join(str(item) for item in value[:5])
    return str(value)


def _format_percent(value: Any) -> str:
    if isinstance(value, (int, float)):
        return f"{value:.2%}" if value <= 1 else f"{value:.2f}%"
    return _format_value(value)
